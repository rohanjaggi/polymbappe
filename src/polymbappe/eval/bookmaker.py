"""Loader for the external bookmaker prediction-accuracy workbook.

The ``WorldCup2026_Prediction_Accuracy*.xlsx`` file is a *shortest-odds favorite* tracker,
not a per-match 1X2 odds book: each row records the bookmaker favorite (shortest odds),
the favorite's moneyline where noted, the score, and whether the favorite outcome
occurred. It therefore supports an honest **accuracy** comparison (favorite pick vs.
model pick, McNemar) but NOT probability-scoring of the market (RPS / log-loss) or CLV,
which need the full home/draw/away price on every match. See
:func:`polymbappe.dashboard.data.bookmaker_comparison`.
"""

from __future__ import annotations

import re
from pathlib import Path

import polars as pl

from polymbappe.data.aliases import normalize_team_name

#: Output schema of :func:`load_bookmaker_accuracy`.
BOOKMAKER_SCHEMA: dict[str, pl.PolarsDataType] = {
    "round": pl.Utf8,
    "fav_team": pl.Utf8,
    "opp_team": pl.Utf8,
    "pair_key": pl.Utf8,
    "book_correct": pl.Boolean,
    "outcome_type": pl.Utf8,
    "fav_moneyline": pl.Int64,
}

_MONEYLINE_RE = re.compile(r"([+-]\d{2,4})")


def _pair_key(a: str, b: str) -> str:
    """Order-independent fixture key from two normalized team names."""

    return " | ".join(sorted((a, b)))


def _parse_moneyline(note: object) -> int | None:
    """Extract the favorite's American moneyline from a free-text note, if present."""

    if note is None:
        return None
    m = _MONEYLINE_RE.search(str(note))
    return int(m.group(1)) if m else None


def default_workbook_path(settings_data_dir: Path) -> Path | None:
    """Locate the accuracy workbook: ``data/raw`` first, then the project root.

    The file ships with an awkward name (``WorldCup2026_Prediction_Accuracy (1).xlsx``);
    matching by glob keeps the loader robust to the copy/rename. Returns ``None`` when no
    workbook is found in either location.
    """

    candidates: list[Path] = []
    raw = settings_data_dir / "raw"
    candidates += sorted(raw.glob("*Prediction_Accuracy*.xlsx"))
    # data_dir is typically ``<root>/data``; also look at the project root.
    candidates += sorted(settings_data_dir.parent.glob("*Prediction_Accuracy*.xlsx"))
    return candidates[0] if candidates else None


def load_bookmaker_accuracy(path: Path) -> pl.DataFrame:
    """Parse the ``Matches`` sheet into a normalized per-fixture accuracy frame.

    Each returned row carries the normalized favorite/opponent names, an order-independent
    ``pair_key`` for joining to model fixtures, whether the bookmaker favorite was correct,
    the outcome type (``Correct`` / ``Draw (fav held)`` / ``Upset``), and the favorite's
    moneyline when the note contained one (else null). Returns a typed empty frame when the
    sheet has no gradable rows.

    Requires ``openpyxl`` (declared in the ``dashboard`` optional-dependencies group).
    """

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - exercised via extras
        raise ModuleNotFoundError(
            "Reading the bookmaker workbook needs openpyxl. Install the dashboard extra: "
            "`pip install -e '.[dashboard]'`."
        ) from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    if "Matches" not in wb.sheetnames:
        return pl.DataFrame(schema=BOOKMAKER_SCHEMA)
    ws = wb["Matches"]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    # Find the header row ("Date", "Round", "Match (Fav vs Opp)", ...).
    header_idx = next(
        (i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Date"), None
    )
    if header_idx is None:
        return pl.DataFrame(schema=BOOKMAKER_SCHEMA)

    records: list[dict[str, object]] = []
    for r in rows[header_idx + 1 :]:
        if not r or r[0] is None:
            continue  # skip blank/footer rows (no date)
        match_label = str(r[2]) if len(r) > 2 and r[2] is not None else ""
        predicted = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        if " v " not in match_label or not predicted:
            continue
        left, right = (s.strip() for s in match_label.split(" v ", 1))
        fav = normalize_team_name(predicted)
        # The opponent is whichever side of the label is not the favorite.
        opp_raw = right if normalize_team_name(left) == fav else left
        opp = normalize_team_name(opp_raw)
        correct = r[7] if len(r) > 7 else None
        records.append(
            {
                "round": str(r[1]) if len(r) > 1 and r[1] is not None else "",
                "fav_team": fav,
                "opp_team": opp,
                "pair_key": _pair_key(fav, opp),
                "book_correct": bool(correct) if correct is not None else False,
                "outcome_type": str(r[8]) if len(r) > 8 and r[8] is not None else "",
                "fav_moneyline": _parse_moneyline(r[4] if len(r) > 4 else None),
            }
        )

    if not records:
        return pl.DataFrame(schema=BOOKMAKER_SCHEMA)
    return pl.DataFrame(records, schema=BOOKMAKER_SCHEMA)
